# -*- coding: utf-8 -*-
# @Time    : 1/17/2025 6:01 PM
# @FileName: sysml2json.py
# @Software: PyCharm
import json
import re
import openpyxl
import os

# 处理文件及保存文件名 ****.xlsx
filename = 'ResponsePlanElement'

# 输入sysml
input_str = ""
with open(filename+'.txt', 'r') as f:  # r为标识符，表示只读
    input_str = f.read()

# 删除空行
input_str = input_str.replace('\n\n', '\n')

'''将sysml转化为json数据'''

def parse_to_json(input_text):
    # 初始化栈和结果字典
    stack = []
    result = {"@type": "package", "@name": "", "children": []}
    current = result

    # 分割输入文本为行
    lines = input_text.strip().split('\n')

    # print(lines)
    for line in lines:
        line = line.strip()
        #print('1', line)
        if 'package' in line:
            _, name = line.split(' ', 1)
            current["@name"] = name[:-1]  # 移除末尾的 '{'
            #print('2', current)
        elif line.endswith('{'):
            # 移除末尾的 '{' 并分割
            parts = line[:-1].strip().split(' ')  # 使用 2 作为分割限制，以正确处理 "part def" 的情况
            #print('3', parts)
            # 创建新的字典并添加到当前的children列表
            type_name = parts[0]
            tar_name = parts[1]
            new_dict = {"@type": type_name, "@name": "", "children": []}
            if len(parts) > 1 and "def" in parts[1]:
                new_dict["@type"] += " " + parts[1]  # 将 'def' 添加到 "@type" 字段
                new_dict["@name"] = parts[2]  # 将名称赋值给 "@name"

            # elif len(parts) > 1 and ':' in parts[2]:
            elif len(parts) > 1 and ':' in parts:
                new_dict = {"@type": type_name, "@name": "", "parent": "", "children": []}
                new_dict["@type"] += "" + "Sub"
                new_dict["@name"] = parts[1]
                new_dict["parent"] = parts[3]

            else:
                new_dict["@name"] = parts[1]

            current["children"].append(new_dict)
            # 将当前字典压入栈并更新当前字典
            stack.append(current)
            current = new_dict
        elif line == '}':
            # 结束当前字典，弹出栈顶元素
            if stack:
                current = stack.pop()

        elif line:
            # 处理没有子元素的行
            parts = line.split(' ')
            # print('qwe',parts)
            for i in range(len(parts)):  # 元素内包含分号，则删除分号，如果不包含分号则不进行任何操作
                if isinstance(parts[i], str) and ';' in parts[i]:
                    parts[i] = parts[i].replace(';', '')

            # print('4', parts)
            if len(parts) == 1:
                current["children"].append({"@type": new_dict["@name"], "@name": parts[0]})

            elif ':' in parts:
                idx = parts.index(':')  # 获取冒号的索引
                name = parts[idx - 1]  # 取冒号前一个元素作为name
                type_parts = ' '.join(parts[:idx - 1])  # 取冒号前的所有元素并保留空格作为type

                # 判断冒号后的元素是instance还是annotation
                if any(keyword in parts[idx + 1] for keyword in ['Boolean', 'String', 'Integer', 'Real']):  # 可添加
                    if idx + 2 < len(parts) and parts[idx + 2] == '=':  # 检查是否存在等号
                        current["children"].append({"@type": type_parts, "@name": name, "datatype": parts[idx + 1],
                                                    "datavalue": ' '.join(parts[idx + 3:]),"owner":tar_name})
                    else:
                        current["children"].append({"@type": type_parts, "@name": name, "datatype": parts[idx + 1],
                                                    "owner":tar_name})

                elif 'ref' in parts and 'part' in parts:
                    current["children"].append(
                        {"@type": 'partAssociate', "@name": ''.join(parts[2]), "datavalue": parts[idx + 1],
                         "owner":tar_name})

                elif 'perform' in parts and 'action' in parts:
                    current["children"].append(
                        {"@type": 'actionSub', "@name": ''.join(name), 'parent': ''.join(parts[idx + 1]),
                         'owner': tar_name})

                elif 'item' in parts:
                    current["children"].append(
                        {"@type": type_parts, "@name": ''.join(name), 'parent': ''.join(parts[idx + 1]),
                         'owner': tar_name})

                else:
                    current["children"].append({"@type": type_parts, "@name": name, "datavalue": parts[idx + 1]})

            # elif 'entry' in parts:
            #     current["children"].append({"@type": parts[0], "@name": ''.join(parts[2:])})

            elif 'part' in parts and 'def' in parts:
                if parts[-1].endswith('{}'):
                    parts[-1] = parts[-1].rstrip('{}')
                current["children"].append({"@type": parts[0] + ' ' + parts[1], "@name": ''.join(parts[2])})
                #print('5')

            elif 'ref' in parts and 'part' in parts:
                current["children"].append({"@type": 'partAssociate', "@name": ''.join(parts[2]), 'owner': tar_name})

            elif 'perform' in parts and 'action' in parts:
                current["children"].append({"@type": 'actionSub', "@name": ''.join(parts[2]), 'owner': tar_name})

            elif 'exhibit' in parts and 'state' in parts:
                #print('6', parts)
                current["children"].append({"@type": 'exhibitState', "@name": ''.join(parts[2]), 'owner': tar_name})

            elif 'redefines' in parts:
                if parts[-2] == '=':  # 检查是否存在等号
                    current["children"].append({"@type": parts[0], "@name": parts[2], "datavalue": ''.join(parts[-1]),
                                                'owner': tar_name})
                else:
                    current["children"].append({"@type": parts[0], "@name": parts[2], 'owner': tar_name})

                #print('7', parts)

            else:
                name = parts[1]
                current["children"].append({"@type": parts[0], "@name": name})
                #print('8')

    return result

# 解析并打印JSON
parsed_json = parse_to_json(input_str)
#print(json.dumps(parsed_json, indent=2, ensure_ascii=False))

'''对json数据进行结构处理'''
def extract_data(data):
    result = []
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, list):
                result.append((key, ''))  # 将键名以元组的形式添加到result中
                for item in value:
                    result.extend(extract_data(item))
            elif isinstance(value, dict):
                result.extend(extract_data(value))
            else:
                result.append((key, value))
    elif isinstance(data, list):
        for item in data:
            result.extend(extract_data(item))
    return result

data_new = extract_data(parsed_json)
#print(data_new)

data_new = extract_data(parsed_json)
#print(data_new)

result = []
temp = {}

for key, value in data_new:
    match = re.match(r'@type', key) # re正则表达式，匹配@type
    if match:
        if temp:
            result.append(temp)
        temp = {}
    temp[key] = value

if temp:
    result.append(temp)

# print(result)
# print(result[1]['@type'])

'''state 处理'''
# state 处理
state_info = []
for i in result:
    if i.get('@type') in ['state def', 'state', 'accept', 'then']:
        state_info.append(i)

# 按照 '@type': 'state def' 分割
split_state_info = []
temp_part = []
for item in state_info:
    if item.get('@type') == 'state def':
        if temp_part:
            split_state_info.append(temp_part)
            temp_part = []
    temp_part.append(item)
if temp_part:
    split_state_info.append(temp_part)

state_list_wh = []
for i in split_state_info:
    state_list = []
    state = {}
    j_1 = []
    for j in i:
        j_1.append(j["@name"])
    #print(j_1)

    state_list.append(j_1[0])

    j_1 = list(dict.fromkeys(j_1))
    #print(j_1)

    for n in range(1,len(j_1)-2,2):
        state_list.append({'source': j_1[n], 'transit': j_1[n+1], 'target': j_1[n+2]})
        #print('1',state_list)

    state_list_wh.append(state_list)


#print(state_list_wh)

#print(state_list_wh)

for i in state_list_wh:
    #print(i)
    for item in result:
        if item.get('@type') == "state def":
            if item.get('@name') == i[0]:
                item["transitions"] = i[1:]
                #print(item)

result = [i for i in result if i.get('@type') not in ['then', 'entry', 'accept', 'state']]
# for i in result:
#     print(i)

'''action 处理'''
# action
# inparams
action_info = []
for i in result:
    if i.get('@type') in ['action def', 'in']:
        action_info.append(i)

#print(action_info)
split_action_info = []
temp_part = []
for item in action_info:
    if item.get('@type') == 'action def':
        if temp_part:
            split_action_info.append(temp_part)
            temp_part = []
    temp_part.append(item)
if temp_part:
    split_action_info.append(temp_part)

#print(split_action_info)

action_list_wh_in = []
for i in split_action_info:
    action_list = []
    action_list_1 = []
    action = {}
    j_1 = []
    for j in i:
        j_1.append(j["@name"])
    #print(j_1)

    action_list_wh_in.append(j_1)

# outparams
action_info = []
for i in result:
    if i.get('@type') in ['action def', 'out']:
        action_info.append(i)

#print(action_info)

split_action_info = []
temp_part = []
for item in action_info:
    if item.get('@type') == 'action def':
        if temp_part:
            split_action_info.append(temp_part)
            temp_part = []
    temp_part.append(item)
if temp_part:
    split_action_info.append(temp_part)

#print(split_action_info)

action_list_wh_out = []
for i in split_action_info:
    action_list = []
    action_list_1 = []
    action = {}
    j_1 = []
    for j in i:
        j_1.append(j["@name"])
    #print(j_1)

    action_list_wh_out.append(j_1)

#print(action_list_wh_out)

for i in action_list_wh_in:
    #print(i)
    for item in result:
        if item.get('@type') == 'action def':
            if item.get('@name') == i[0]:
                item["inparams"] = i[1:]
                #print(item)

for i in action_list_wh_out:
    #print(i)
    for item in result:
        if item.get('@type') == 'action def':
            if item.get('@name') == i[0]:
                item["outparams"] = i[1:]
                #print(item)

'''转换名称'''
for item in result:
    if item.get('@type') == 'part def':
        item['@type'] = 'part'

for item in result:
    if item.get('@type') == 'item':
        item['@type'] = 'itemComposition'

for item in result:
    if item.get('@type') == 'item def':
        item['@type'] = 'item'

for item in result:
    if item.get('@type') == 'attribute def':
        item['@type'] = 'attribute'

for item in result:
    if item.get('@type') == 'action def':
        item['@type'] = 'action'

for item in result:
    if item.get('@type') == 'state def':
        item['@type'] = 'state'

result = [i for i in result if i.get('@type') not in ['in', 'out']]
# for i in result:
#     print(i)

'''attribute redefine 处理'''
datatype_mapping = {}
for i in result:
    if i['@type'] == 'attribute' and 'datatype' in i:
        datatype_mapping[i['@name']] = i['datatype']
        #print(datatype_mapping)

# 遍历数据，为缺少datatype的attribute添加datatype
for i in result:
    if i['@type'] == 'attribute' and 'datatype' not in i:
        i['datatype'] = datatype_mapping.get(i['@name'], 'none')

'''attribute 处理'''
attribute_data = [item for item in result if item.get('@type') == 'attribute']
attribute_dict = {}
for item in attribute_data:
    name = item['@name']
    owner = item['owner']
    if name not in attribute_dict:
        attribute_dict[name] = item
    elif owner != 'def' and attribute_dict[name]['owner'] == 'def':
        attribute_dict[name] = item

for name, item in attribute_dict.items():
    if item['owner'] == 'def':
        item['owner'] = 'none'

result = [item for item in result if item.get('@type') != 'attribute'] + list(attribute_dict.values())

'''partAssociate 处理'''
partass_data = [item for item in result if item.get('@type') == 'partAssociate']
partass_dict = {}
for item in partass_data:
    name = item['@name']
    owner = item['owner']
    if name not in partass_dict:
        partass_dict[name] = item
    elif owner != 'def' and partass_dict[name]['owner'] == 'def':
        partass_dict[name] = item

for name, item in partass_dict.items():
    if item['owner'] == 'def':
        item['owner'] = 'none'

result = [item for item in result if item.get('@type') != 'partAssociate'] + list(partass_dict.values())

for i in result:
    print(i)

'''保存数据'''
result_json = json.dumps(result)

with open(filename+'.json', 'w') as file:
    file.write(result_json)
    print("成功保存：" + filename + '.json')

'''tpye 与 name 对应表'''
type_name_dict = {}

for item in result:
    item_type = item["@type"]
    item_name = item["@name"]

    if item_type in type_name_dict:
        type_name_dict[item_type].append(item_name)
    else:
        type_name_dict[item_type] = [item_name]

print("\ntpye name 对应表")
for item_type, item_names in type_name_dict.items():
    print(f"type:'{item_type}'------>name:{item_names}")

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "type"
sheet["B1"] = "name"
sheet["C1"] = "children"
sheet["D1"] = "parent"
sheet["E1"] = "datatype"
sheet["F1"] = "datavalue"
sheet["G1"] = "owner"
sheet["H1"] = "inparams"
sheet["I1"] = "outparams"
sheet["J1"] = "transitions.source"
sheet["K1"] = "transitions.transit"
sheet["L1"] = "transitions.target"

# 写入数据
row_num = 2
for item in result:
    if "transitions" in item and item["transitions"]:
        for i in item["transitions"]:
            sheet.cell(row=row_num, column=1, value=item["@type"])
            sheet.cell(row=row_num, column=2, value=item["@name"])

            if "children" in item and item["children"]:
                sheet.cell(row=row_num, column=3, value=item["children"])

            if "parent" in item and item["parent"]:
                sheet.cell(row=row_num, column=4, value=item["parent"])

            if "datatype" in item and item["datatype"]:
                sheet.cell(row=row_num, column=5, value=item["datatype"])

            if "datavalue" in item and item["datavalue"]:
                sheet.cell(row=row_num, column=6, value=item["datavalue"])

            if "owner" in item and item["owner"]:
                sheet.cell(row=row_num, column=7, value=item["owner"])
            sheet.cell(row=row_num, column=10, value=i["source"])
            sheet.cell(row=row_num, column=11, value=i["transit"])
            sheet.cell(row=row_num, column=12, value=i["target"])
            row_num += 1
    num = 0
    if "inparams" in item and item["inparams"]:
        for num in range(len(item["inparams"])):
            print(num)
            sheet.cell(row=row_num, column=1, value=item["@type"])
            sheet.cell(row=row_num, column=2, value=item["@name"])

            if "children" in item and item["children"]:
                sheet.cell(row=row_num, column=3, value=item["children"])

            if "parent" in item and item["parent"]:
                sheet.cell(row=row_num, column=4, value=item["parent"])

            if "datatype" in item and item["datatype"]:
                sheet.cell(row=row_num, column=5, value=item["datatype"])

            if "datavalue" in item and item["datavalue"]:
                sheet.cell(row=row_num, column=6, value=item["datavalue"])

            if "owner" in item and item["owner"]:
                sheet.cell(row=row_num, column=7, value=item["owner"])

            sheet.cell(row=row_num, column=8, value=item["inparams"][num])

            if item["outparams"]:
                sheet.cell(row=row_num, column=9, value=item["outparams"][num])
            row_num += 1

        if num < len(item["outparams"]):
            print(123213, i)
            for num in range(num + 1, len(item["outparams"])):
                print(i)
                sheet.cell(row=row_num, column=1, value=item["@type"])
                sheet.cell(row=row_num, column=2, value=item["@name"])

                if "children" in item and item["children"]:
                    sheet.cell(row=row_num, column=3, value=item["children"])

                if "parent" in item and item["parent"]:
                    sheet.cell(row=row_num, column=4, value=item["parent"])

                if "datatype" in item and item["datatype"]:
                    sheet.cell(row=row_num, column=5, value=item["datatype"])

                if "datavalue" in item and item["datavalue"]:
                    sheet.cell(row=row_num, column=6, value=item["datavalue"])

                if "owner" in item and item["owner"]:
                    sheet.cell(row=row_num, column=7, value=item["owner"])

                sheet.cell(row=row_num, column=9, value=item["outparams"][num])
                row_num += 1
    else:
        sheet.cell(row=row_num, column=1, value=item["@type"])
        sheet.cell(row=row_num, column=2, value=item["@name"])

        if "children" in item and item["children"]:
            sheet.cell(row=row_num, column=3, value=item["children"])

        if "parent" in item and item["parent"]:
            sheet.cell(row=row_num, column=4, value=item["parent"])

        if "datatype" in item and item["datatype"]:
            sheet.cell(row=row_num, column=5, value=item["datatype"])

        if "datavalue" in item and item["datavalue"]:
            sheet.cell(row=row_num, column=6, value=item["datavalue"])

        if "owner" in item and item["owner"]:
            sheet.cell(row=row_num, column=7, value=item["owner"])
        row_num += 1

    # row_num += 1

# 保存文件
workbook.save(filename + '_1.xlsx')

if os.path.exists(filename + '_1.xlsx'):
    print('成功保存：' + filename + '_1.xlsx')
else:
    print('保存失败')